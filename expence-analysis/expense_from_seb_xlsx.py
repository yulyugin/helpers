"""
    expence_from_seb_xlsx.py - get expense data from XLSX file producesd
    by SEB bank.

    Copyright (C) 2017 Evgenii Iuliugin

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""

import openpyxl
from datetime import datetime
from itertools import islice
import re

from transaction import Transaction

def expense_reader(filename):
    ret = list()
    wb = openpyxl.load_workbook(filename = filename)

    # Bokföringsdatum, Valutadatum, Verifikationsnummer, Text, Belopp, Saldo
    receiver_column = 3
    date_column  = 1
    amount_column = 4
    if wb.sheetnames[0] == 'ExportSida1':
        rows_to_skip = 5
    elif wb.sheetnames[0] == 'Sheet1':
        rows_to_skip = 8
    else:
        assert False, f'Unknown SEB format. First sheet: {wb.sheetnames[0]}'

    p = wb[wb.sheetnames[0]]
    # Skip lines that doesn't contatin relevant transaction information
    for row in islice(p.iter_rows(), rows_to_skip, None):
        # Remove trailing spaces
        receiver = row[receiver_column].value.strip()
        date = None

        # Receiver is usually in "Name/YY-MM-DD" format.
        # Separate these fields
        namedate = receiver.split('/')
        if len(namedate) >= 2:
            try:
                date = datetime.strptime(namedate[-1], "%y-%m-%d").date()
            except ValueError:
                pass
            else:
                # '/' symbols can be in the middle of the string
                receiver = receiver[:-(len(namedate[-1]) + 1)].strip()

        if date == None:
            # Use date_column if date is not available in receiver field
            date = datetime.strptime(row[date_column].value, "%Y-%m-%d").date()

        # Remove special symbols from the end of the name
        receiver = re.sub(r'(([^\w]|_|\s)+)$', '', receiver)

        ret.append(Transaction(date, receiver, row[amount_column].value))

    return ret
